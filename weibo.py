import openpyxl
import requests
import time
import re
import urllib.parse
from bs4 import BeautifulSoup
import random
import jieba
import pandas as pd
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def word_frequency_xlsx(xlsx_name='weibo_comment.xlsx'):
    df = pd.read_excel(xlsx_name)  # 读取 Excel 文件
    txt_name = "xlsx_to_txt.txt"
    m = 1
    # with open(txt_name, 'w', encoding="utf-8") as fp:  # 清空文件
    #     pass

    while m < len(df):
        data = df.iloc[m, 0]  # 直接获取数据
        with open(txt_name, 'a', encoding="utf-8") as fp:  # 追加模式
            fp.writelines(str(data) + '\n')  # 每行数据换行
        m += 1  # 递增索引
    word_frequency_txt(txt_name)


def word_frequency_txt(txt_name='weibo_comment.txt'):
    """
    统计词频
    """
    with open('baidu_stopwords.txt', 'r', encoding='utf-8') as f:   #读入停用词文件
        stopwords = set(line.strip() for line in f)

    # 读入文件
    with open(txt_name, encoding="utf-8") as f:
        text = f.read()

    ls = jieba.lcut(text, cut_all=True)  # 分词
    # 统计词频
    counts = {}
    for i in ls:
        if len(i) > 1:
            counts[i] = counts.get(i, 0) + 1
    for word in stopwords:  # 去掉停用词
        counts.pop(word, 0)
    ls1 = sorted(counts.items(), key=lambda x: x[1], reverse=True)  # 词频排序
    # 输出
    print(ls1)
    with open('word_frequency.txt', 'w', encoding="utf-8") as fp:
        fp.writelines(str(ls1) + '\n')
def save_to_txt(comment_list, txt_name='weibo_comment.txt'):
    """
    保存评论到txt，不保存地址和点赞数
    """
    with open(txt_name, 'w', encoding="utf-8") as f:  # 清空文件
        pass
    with open(txt_name, 'a', encoding='utf-8') as f:  # 追加模式
        for comment in comment_list:
            f.write(comment + '\n')     # 每行数据换行

def save_to_excel(comment_list, location_list, like_list, xls_name='weibo_comment.xlsx'):
    """
    保存数据到excel
    """
    book = openpyxl.Workbook()
    sheet = book.create_sheet('微博评论', index=0)

    sheet.cell(1, 1).value = '评论'
    sheet.cell(1, 2).value = '地址'
    sheet.cell(1, 3).value = '点赞数'
    row = 2
    for i in range(len(comment_list)):
        sheet.cell(row, 1).value = comment_list[i]
        sheet.cell(row, 2).value = location_list[i]
        sheet.cell(row, 3).value = like_list[i]
        row += 1
    book.save(xls_name)



def get_son_comment(max_id2, id2, uid2):
    """
    获取子评论数据
    :param max_id2: 用于翻页
    :param id2: 主评论ID
    :param uid2: 用户ID
    :return: max_id2
    """
    url = f"https://weibo.com/ajax/statuses/buildComments?flow=0&is_reload=1&id={id2}&is_show_bulletin=2&is_mix=1&fetch_level=1&max_id= {max_id2}&count=20&uid={uid2}&locale=zh-CN"
    rep3 = request_weibo(url)
    rep3.encoding = "utf-8"
    comment2 = rep3.json()
    data2 = comment2['data']
    for i in range(len(data2)):
        comment = ILLEGAL_CHARACTERS_RE.sub(r'', data2[i]['text_raw'])
        comment_list.append(comment)

        position = ILLEGAL_CHARACTERS_RE.sub('r', data2[i]['user']['location'])
        location_list.append(position)

        like = data2[i]['like_counts']
        like_list.append(str(like))

    max_id2 = str(comment2['max_id'])
    rep3.close()
    return max_id2

def get_comment(count, max_id1, id1, uid1):
    """
    获取主评论列表，并递归获取子评论
    :param count: 每页评论数
    :param max_id1: 当前页面的最大ID（用于翻页）
    :param id1: 该微博的ID
    :param uid1: 该微博作者的用户ID
    :return: max_id, 总评论数
    """
    url = f"https://weibo.com/ajax/statuses/buildComments?is_reload=1&id={id1}&is_show_bulletin=2&is_mix=0&max_id={max_id1}&count={count}&uid={uid1}&fetch_level=0&locale=zh-CN"
    rep2 = request_weibo(url)
    rep2.encoding = "utf-8"
    comments = rep2.json()
    data = comments['data']
    for i in range(len(data)):
        comment = ILLEGAL_CHARACTERS_RE.sub(r'', data[i]['text_raw'])
        comment_list.append(comment)

        position = ILLEGAL_CHARACTERS_RE.sub('r', data[i]['user']['location'])
        location_list.append(position)

        like = data[i]['like_counts']
        like_list.append(str(like))

        total_number = data[i]['total_number']
        page_num = int(total_number / 20)
        if total_number != 0:
            id2 = data[i]['id']
            uid2 = data[i]['user']['id']
            max_id2 = "0"
            max_id2_list = ["0"]
            for i in range(0, page_num + 2):
                max_id2 = get_son_comment(max_id2, id2, uid2)
                if max_id2 not in max_id2_list:
                    max_id2_list.append(max_id2)
                    continue
                else:
                    break

    max_id = str(comments['max_id'])
    rep2.close()
    return max_id, int(comments['total_number'])


def request_weibo(search_url):
    """
    嘗試訪問網站
    """
    n = 0
    time.sleep(random.random())
    while True:
        try:
            response = requests.get(search_url, headers=headers)
            n += 1
            if response.status_code == 200:
                print(search_url + "访问成功")
                return response
            else:
                n += 1
                continue
        except requests.exceptions.RequestException as e:
            if n > 20:
                print("尝试访问网站超过20次，任未成功，请确认输入是否正确\n")
                print(search_url)
                print(e)
                break
            continue

def get_url_mid_id(search_url):
    """
    獲取搜索網站關鍵詞裏的博客url、mid、id(uid)
    """
    mid_list = []
    uid_list = []
    urls_list = []

    html = request_weibo(search_url)
    soup = BeautifulSoup(html.text, 'lxml')
    # print(soup)
    div_from = soup.find_all('div', class_='from')
    div_card_wrap = soup.find('div', class_='main-full').find_all('div', class_='card-wrap')
    # 找出各个博客url
    for div in div_from:
        urls_list.append("https:" + div.a['href'])
    # 找出各个博客mid
    num = 0
    for div in div_card_wrap:
        try:
            mid_list.append(div_card_wrap[num]['mid'])
            num += 1
        except:
            num += 1
            continue
    # 用正则表达式解析出uid
    pattern = r"weibo\.com/(\d+)"
    for url in urls_list:
        uid_list.append(re.search(pattern, url).group(1))
    return mid_list, uid_list, urls_list
    # print(urls_list)
    # print(mid_list)
    # print(uid_list)


def text_to_encoded(keyword_list):
    """
    将关键词转换为URL编码（好像不用也行）
    """
    encoded_list = []
    for keyword in keyword_list:
        encoded_list.append(urllib.parse.quote(keyword))
    return encoded_list


def loop_get_comment(encoded, page=50):
    """
    循环获取“page”页数据
    """
    global total_num
    search_url = 'https://s.weibo.com/weibo?q=' + str(encoded)
    for page in range(1, page):
        [mid_list, uid_list, urls_list] = get_url_mid_id(search_url + "&page=" + str(page))
        # 基本参数count、max_id
        count = '10'
        max_id = '0'
        # 得到博客的評論
        for i in range(len(urls_list)):
            id1 = mid_list[i]
            uid1 = uid_list[i]
            max_id_list = ["0"]
            while True:
                max_id, num = get_comment(count, max_id, id1, uid1)
                count = '20'
                if max_id not in max_id_list:
                    max_id_list.append(max_id)
                    continue
                else:
                    total_num += num
                    break




if __name__ == '__main__':
    headers = {
        "Referer": "https://weibo.com/",  # 用于告诉服务器请求来源的页面(需要则修改
        "Cookie": "更换新Cookie",
        # 报错则更新
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        # 可不动
    }
    comment_list = []
    location_list = []
    like_list = []
    total_num = 0

    keyword_list = ['新能源汽车']
    encoded_list = text_to_encoded(keyword_list)

    page = 50           # 页数为2-50页，其他页数都是抓取第一页
    for encoded in encoded_list:
        loop_get_comment(encoded, page)
    print("微博显示的总数据量" + str(total_num))    # 微博显示的总数据量
    print("去除屏蔽、删除后实际能抓取的数据量" + str(len(comment_list)))    # 去除屏蔽、删除后实际能抓取的数据量
    # 保存于excel
    save_to_excel(comment_list, location_list, like_list)
    word_frequency_xlsx()


    # # 保存于txt，并统计词频
    # save_to_txt(comment_list)
    # word_frequency_txt()

    # txt_name = 'weibo_comment.txt'   # 默认为weibo_comment.txt
    # save_to_txt(comment_list, txt_name=txt_name)
    # word_frequency_txt(txt_name=txt_name)
